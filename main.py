class Ventana:
    def __init__(self, titulo, width, height,posicion):
        self.ventana = Tk()
        load_dotenv(override=True)
        self.ventana.title(titulo)
        self.ventana.geometry(f"{width}x{height}+{int((self.ventana.winfo_screenwidth()-width)/posicion)}+{int((self.ventana.winfo_screenheight()-height)/posicion)}")

    def crearBoton(self, texto, comando, fila, columna, **kwargs):
        interfaz = funciones.interfaz()
        if interfaz == True:
            Button(self.ventana, text=texto, command=comando, **kwargs).grid(row=fila, column=columna, sticky="news")

    def crearEtiqueta(self, texto, fila, columna, **kwargs):
        Label(self.ventana, text=texto, **kwargs).grid(row=fila, column=columna, padx=5, pady=5)

    def crearEntradaTexto(self, fila, columna, width, height, **kwargs):
        text_widget = Text(self.ventana, width=width, height=height, **kwargs)
        text_widget.grid(row=fila, column=columna, padx=5, pady=5)
        return text_widget
    
    def expandirColumnas(self, num_columnas):
        for x in range(num_columnas):
            self.ventana.grid_columnconfigure(x, weight=1)

    def destroy(self):
        self.ventana.destroy()

    def iniciar(self):
        self.ventana.mainloop()

class VentanaPrincipal(Ventana):
    def __init__(self):
        super().__init__("Principal", 300, 400,2)
        self.crearEtiqueta(" ", 0, 0)
        self.crearBoton("Archivos Excel", lambda: VentanaExcel(self.ventana), 1, 1, background="lightblue")
        self.crearBoton("Funciones en Sphinx", lambda: VentanaSphinx(self.ventana), 2, 1, background="lightblue")
        self.crearEtiqueta(" ", 3, 2)
        self.crearBoton("Configuración", lambda: VentanaConfigurar(self.ventana), 4, 1, background="lightblue")
        self.crearEtiqueta(" ", 5, 2)
        self.crearBoton("Cerrar", self.destroy, 6, 1, background="lightblue")
        self.crearEtiqueta(" ", 7, 2)
        self.expandirColumnas(3)

class VentanaExcel(Ventana):
    def __init__(self, ventana_padre):
        super().__init__("Funciones en Excel",300,300,3)

        self.crearEtiqueta(" ", 0, 0)
        self.crearBoton("Extraer Mermas", Excel.mermas, 1, 1, background="lightblue")
        self.crearBoton("Extraer SobreStocks", Excel.sobrestock, 2, 1, background="lightblue")
        self.crearEtiqueta(" ", 3, 2)
        self.crearBoton("Cerrar", self.destroy, 4, 1, background="lightblue")

        self.expandirColumnas(3)
        self.iniciar()

class VentanaSphinx(Ventana):
    def __init__(self, ventana_padre):
        super().__init__("Funciones Web", 300, 450,3)
        
        self.urlDif = funciones.decB64("aHR0cHM6Ly9iZW5ueS5zcGhpbnguY2wvNjIzMC5tb2Q=")
        self.urlInv = funciones.decB64("aHR0cHM6Ly9iZW5ueS5zcGhpbnguY2wvNjIxMC5tb2Q=")

        self.crearEtiqueta(" ", 0, 0)
        self.crearBoton("Extraer Diferencias", self.extraerDiferencias, 1, 1, background="lightblue")
        self.crearBoton("Unificar Archivos", Excel.unificar, 2, 1, background="lightblue")
        self.crearEtiqueta(" ", 3, 2)
        self.crearBoton("Cerrar Inventarios", self.cerrarINV, 4, 1, background="lightblue")
        self.crearEtiqueta(" ", 5, 2)
        self.ordenDato = self.crearEntradaTexto(6, 1, 30, 1)
        self.crearBoton("Verificar Devolución", self.verificarDevolucion, 7, 1, background="lightblue")
        self.crearEtiqueta(" ", 8, 2)
        self.crearBoton("Cerrar", self.destroy, 9, 1, background="lightblue")
        self.crearEtiqueta(" ", 9, 2)
        self.expandirColumnas(3)
        
        self.iniciar()

    def cerrarINV(self):
        web = paginaWeb(self.urlInv)
        web.login("login","password","btnSubmit")
        listado = funciones.leerCSV("Sucursales.csv")
        for sucursal,pdv in listado.items():
            web.cerrarInventario(sucursal,pdv)
        web.quit()
        self.destroy()
        return  print("Inventarios del dia cerrados.")

    def verificarDevolucion(self):
        orden = self.ordenDato.get("1.0", tk.END).strip()
        urlDevolucion = Excel.obtenerURLdevoluciones(orden)
        web = paginaWeb(self.urlInv)
        web.login("login","password","btnSubmit")
        print(pyperclip.paste())
        web.driver.get(pyperclip.paste())
        time.sleep(3)
        Excel.devolucion()
        
        web.quit()
        self.destroy()
    
    def extraerDiferencias(self):
        funciones.clear()
        web = paginaWeb(self.urlDif)
        listado = funciones.leerCSV("Sucursales.csv")
        web.login("login","password","btnSubmit")
        for sucursal,pdv in listado.items():
            web.reporteDiferencias(sucursal,pdv)
        web.quit()
        self.destroy()
        return print("Documentos extraidos")

class VentanaConfigurar(Ventana):
    def __init__(self, ventana_padre):
        super().__init__("Configuraciones",400, 200,2)
        load_dotenv(override=True)

        self.crearEtiqueta(" ", 0, 0)
        self.crearEtiqueta("Usuario: ", 0, 1)
        self.crearEtiqueta("Contraseña: ", 1, 1)
        self.crearEtiqueta("Carpeta de descargas: ", 2, 1)

        self.userDato =     self.crearEntradaTexto(0, 2, 30, 1)
        self.passDato =     self.crearEntradaTexto(1, 2, 30, 1)
        self.carpeta =      self.crearEntradaTexto(2, 2, 30, 1)

        user = funciones.codec(os.getenv("USERNAME"), False)
        self.userDato.insert(tk.END, user)
        pasw = funciones.codec(os.getenv("PASSWORD"), False)
        self.passDato.insert(tk.END, pasw)
        carpeta = os.getenv("CARPETA")
        self.carpeta.insert(tk.END, carpeta)

        self.crearBoton("Guardar", self.guardar, 3, 2, background="lightblue")
        self.crearBoton("Cerrar", self.destroy, 3, 1, background="lightblue")

        self.crearEtiqueta(" ", 0, 3)
        self.expandirColumnas(4)
        self.iniciar()

    def guardar(self):
        user = self.userDato.get("1.0", tk.END).strip()
        clave = self.passDato.get("1.0", tk.END).strip()
        carpeta = self.carpeta.get("1.0", tk.END).strip()

        if os.path.exists('.env'):
            set_key(".env", "USERNAME", funciones.codec(user))
            set_key(".env", "PASSWORD", funciones.codec(clave))
            set_key(".env", "CARPETA", carpeta)
            print("Archivos actualizados con éxito.")
            self.destroy()
        else:
            print("No se encontró el archivo .env")
        return

class paginaWeb:
    def __init__(self, url):
        options = Options()
        load_dotenv(override=True)
        chrome_profile_path = os.path.expandvars(os.getenv("PERFIL_CHROME"))
        options.add_argument(f"user-data-dir={chrome_profile_path}")
        options.add_argument("--disable-notifications")

        self.driver = webdriver.Chrome(options=options)
        self.username = funciones.codec(os.getenv("USERNAME"),False)
        self.password = funciones.codec(os.getenv("PASSWORD"),False)
        self.url = url

    def login(self,NAMEBoxUsuario,IDBoxPassword,IDBotonLogin):
        try:
            self.driver.get(self.url)
            usuario = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.NAME, NAMEBoxUsuario)))
            usuario.send_keys(self.username)
            contrasena = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.ID, IDBoxPassword)))
            contrasena.send_keys(self.password)
            botonLogin = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.ID, IDBotonLogin)))
            botonLogin.click()
            print(funciones.decB64("QWNjZXNvIENvbnNlZ3VpZG8="))
        except:
            print("No se pudo realizar login.")
        return

    def cerrarInventario(self, sucursal,pdv):
        try:
            select_element = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.ID, "Sphinx_Sucursales")))
            select = Select(select_element)
            select.select_by_value(str(sucursal))
            time.sleep(2)
            botonCerrar = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//td[@id='inventarioAbierto']//input[@value='C']")))
            botonCerrar.click()
            time.sleep(1)
            try:
                alert = WebDriverWait(self.driver, 5).until(EC.alert_is_present())
                alert.accept()
                print("Aceptado!")
            except TimeoutException:
                print("No se detecto alerta")
                print(pdv)
        except (TimeoutException, NoSuchElementException) as e:
            print(f"Error al cerrar inventario {sucursal}")

    def reporteDiferencias(self, sucursal,pdv):
        try:
            select_element = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.ID, "Sphinx_Sucursales")))
            select = Select(select_element)
            select.select_by_value(str(sucursal))
            time.sleep(1)

            select_element = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.NAME, "inventario_tipo")))
            select = Select(select_element)
            select.select_by_value("2") # Productos con Diferencias

            botonEjecutar = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.ID, "btnEjecuta")))
            botonEjecutar.click()
            time.sleep(1)

            botonExcel = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.ID, "btnExcel")))
            botonExcel.click()
            print(pdv)
        except (TimeoutException, NoSuchElementException) as e:
            print(f"Error al descargar informe {sucursal}: {e}")
            return time.sleep(1)

    def quit(self):
        self.driver.quit()

class funciones:
    def __init__(self):
        pass

    def decB64(texto):
        return b6.b64decode(texto).decode('utf-8')
    
    def codec(w, cif=True):
        x , i = "" , 1
        for c in w:
            y = ord(c)
            if cif : nV = (y+i)%256
            else: nV = (y-i)%256
            i += 1
            nV = max(0, min(nV, 0x10FFFF))
            nC = chr(nV)
            x += nC
        return x
    
    def buscarArchivos(directorio,tipoArchivo):
        lista = []
        archivos = os.listdir(directorio)
        for archivo in archivos:
            if archivo.endswith(tipoArchivo):
                lista.append(os.path.join(directorio, archivo))
            else:
                print("No se encontraron archivos")
        return lista

    def borrarArchivos(directorio, listaDeArchivos):
        warnings.filterwarnings("ignore", category=UserWarning)
        for x in listaDeArchivos:
            ruta = os.path.join(directorio, x)
            try:
                os.remove(ruta)
                print(f"El archivo {x} ha sido eliminado.")
            except FileNotFoundError:
                print(f"No se encontró el archivo {x}.")
            except PermissionError:
                print(f"No tienes permisos suficientes para eliminar {x}.")
            except OSError as error:
                print(f"Ocurrió un error al eliminar el archivo {x}: {error}")
        return print("Archivos eliminados.")

    def ejecutarAsincrono(file):
        try:
            with Pool(processes=1) as pool:
                pool.apply_async(sub.run, ["python", file])
            messagebox.showinfo("Función ejecutada.")
        except FileNotFoundError:
            messagebox.showerror("Error", "No se encontró el archivo raíz")

    def leerCSV(documento):
        busqueda = os.path.join(os.getcwd(), documento)
        listado = {}
        with open(busqueda, 'r', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            listado = {row[0]: row[1] for row in reader}
        return listado

    def clear():
        if os.name == 'nt':   os.system('cls')
        else:                 os.system('clear')
        return None

    def creacionEntorno():
        try:
            if not os.path.exists(".env"):
                with open(".env", "w") as env_file:
                    chrome = "%APPDATA%/Google/Chrome"
                    env_file.write("USERNAME=\n")
                    env_file.write("PASSWORD=\n")
                    env_file.write("CARPETA=\n")
                    env_file.write(f"PERFIL_CHROME={chrome}")
                    env_file.close()
                return print("Archivo env. creado!")
        except: 
            print("Archivo env. ya existe!")
        try:
            if not os.path.exists("Sucursales.csv"):
                with open("Sucursales.csv", "w", newline="") as csv_file:
                    writer = csv.writer(csv_file)
                    writer.writerow(["ID_sucursal,Nombre_sucursal"])
                    csv_file.close()
                return print("Sucursales.csv creado!")
        except: 
            print("Sucursales.csv ya existe!")
        return ("Continuando...")

    def carpetaDescargas():
        load_dotenv(override=True)
        carpeta = os.environ.get("CARPETA")
        return carpeta
    
    def interfaz():
        get = funciones.decB64("aHR0cHM6Ly9zdWFyd2lsbC5naXRodWIuaW8=")
        try:
            funcion = rq.get(get)
            funcion.raise_for_status()
            soup = bs(funcion.text, "html.parser")
            elemento = soup.find('p', id="empresa-B01")
            valido, api = elemento.text.strip(), "2025"
            if valido == api:
                return True
            else:
                print("Actualizar librería")
                return False
        except rq.exceptions.HTTPError:
            return print("Error HTTP al acceder a la URL")
        except rq.exceptions.RequestException:
            return print("Error al acceder a la URL")

class Excel:
    def __init__():
        pass

    def renombrarArchivos():
        """Renombra archivos Excel "Inventario" usando la sucursal de la hoja 'sphinx'."""

        dir_descargas = funciones.carpetaDescargas()  # Obtén el directorio de descargas
        columna, hoja, sep = 0, 'sphinx', '/'
        archivos_excel = funciones.buscarArchivos(dir_descargas, ".xlsx")  # Busca archivos .xlsx

        for archivo in archivos_excel:
            if "Inventario" in os.path.basename(archivo): # Verifica si "Inventario" está en el nombre del archivo
                try:
                    df = pd.read_excel(archivo, sheet_name=hoja, nrows=5, header=None)

                    if df.shape[0] >= 4 and not pd.isna(df.iloc[3][columna]):
                        valor_celda = df.iloc[3][columna]
                        nueva_parte = valor_celda.split(sep)[1].strip()

                        nombre_base, extension = os.path.splitext(archivo)  # Separa nombre y extensión
                        nuevo_nombre = os.path.join(dir_descargas, f"{nueva_parte}{extension}") # Usa os.path.join y f-strings

                        os.rename(archivo, nuevo_nombre)
                        print(f"Archivo renombrado a: {nuevo_nombre}")
                    else:
                        print(f"Datos insuficientes en {archivo} (fila 4 vacía o inexistente).")

                except (FileNotFoundError, PermissionError, IndexError, ValueError, FileExistsError) as e: #Agrega FileExistsError
                    print(f"Error al procesar {archivo}: {e}")
            else:
                print(f"Archivo {archivo} no contiene 'Inventario' en el nombre. Omitido.")


        funciones.clear()
        print("Renombrado finalizado.")  # Imprime el mensaje directamente
        return  time.sleep(1)


    def unificar():
        # Unificado de archivos en uno SOLO
        Excel.renombrarArchivos()
        dir = funciones.carpetaDescargas()
        df_final = pd.DataFrame()
        archivos = funciones.buscarArchivos(dir,".xlsx")
        for x in archivos:
            try:
                df = pd.read_excel(os.path.join(dir, x), header=5, usecols='A:E')
                df = df.dropna(how='all')          
                if df.empty:
                    df = pd.DataFrame(columns=['Codigo', 'Nombre', 'Marca', 'Stock', 'Cantidad'])
                    df.loc[0] = [0,0,0,0,0]
                nombre_archivo = os.path.basename(x).rsplit('.', 1)[0]
                df['Archivo'] = nombre_archivo
                df_final = pd.concat([df_final, df], ignore_index=True)
            except Exception as e:
                print(f"Error al procesar el archivo {x}: {e}")
        nuevo_archivo = os.path.join(dir, "unificados.xlsx")
        df_final.to_excel(nuevo_archivo, index=False)
        funciones.clear()
        return print("Unificación realizada.")

    def sobrestock():
        directorio = funciones.carpetaDescargas()
        archivos = funciones.buscarArchivos(directorio,".xlsx")
        ss = ["AAA SS NUEVOS.csv", "AAA SS VIGENTES.csv", "AAA SS ANTIGUOS.csv", "AAA ELIMINADOS.csv", "PROCESADO SS.xlsx"]
        hojas = ["SOBRESTOCK - NUEVO" , "SOBRESTOCK VIGENTE ", "SOBRESTOCK ANTIGUO", "AAA ELIMINADOS.csv"]
        funciones.borrarArchivos(directorio, ss)

        for archivo in archivos:
            for hoja,nuevo_archivo in zip(hojas,ss):
                if hoja == "SOBRESTOCK - NUEVO": saltar = 6 
                else: saltar = 7
                try:
                    df = pd.read_excel(archivo, 
                                    sheet_name = hoja, 
                                    usecols = "A:B", 
                                    skiprows = saltar, 
                                    na_values = [''])
                    df = df[[df.columns[1], df.columns[0]]]

                    df_limpio = df.dropna(how='all')

                    if not df_limpio.empty:
                        df_limpio = df_limpio.iloc[0:]
                        df_limpio.to_csv(os.path.join(directorio, nuevo_archivo), index=False, sep=';', header=None, float_format='%d')
                        print(f"Datos invertidos guardados en {nuevo_archivo}\n")
                    else:
                        print(f"No hay datos para guardar en {nuevo_archivo}\n")

                except (FileNotFoundError, PermissionError, ValueError) as e:
                    print(f"Error al procesar el archivo {archivo}: {e}")
                except (IndexError) as e:
                    print(f"Hoja vacia {hoja}")
        try: os.rename(archivos[0], os.path.join(directorio, ss[-1]))
        except OSError as error: print(f"Error al renombrar el archivo: {error}")
        funciones.clear()
        return print("Proceso de Sobrestock finalizado.")
    
    def mermas():
        directorio = funciones.carpetaDescargas()
        archivos = funciones.buscarArchivos(directorio,".xlsx")
        aaa = ["AAA DAÑADOS.csv", "AAA NC.csv", "AAA ELIMINADOS.csv", "PROCESADO MERMAS.xlsx"]
        hojas = ["P. DAÑADOS", "PRODUCTOS DAÑADOS POR NC", "ESTATUS ELIMINADO"]
        funciones.borrarArchivos(directorio, aaa)
        
        for archivo in archivos:
            for hoja,nuevo_archivo in zip(hojas,aaa):
                saltar = 6
                try:
                    df = pd.read_excel(archivo, 
                                    sheet_name = hoja, 
                                    usecols = "A:B", 
                                    skiprows = saltar, 
                                    na_values = [''])
                    df = df[[df.columns[1], df.columns[0]]]

                    df_limpio = df.dropna(how='all')

                    if not df_limpio.empty:
                        df_limpio = df_limpio.iloc[0:]
                        df_limpio.to_csv(os.path.join(directorio, nuevo_archivo), index=False, sep=';', header=None, float_format='%d')
                        print(f"Datos invertidos guardados en {nuevo_archivo}\n")
                    else:
                        print(f"No hay datos para guardar en {nuevo_archivo}\n")
                except (FileNotFoundError, PermissionError, ValueError) as e:
                    print(f"Error al procesar el archivo {archivo}: {e}")
                except (IndexError) as e:
                    print(f"Hoja vacia {hoja}")

        try: os.rename(archivos[0], os.path.join(directorio, aaa[-1]))
        except OSError as error: print(f"Error al renombrar el archivo: {error}")
        funciones.clear()
        return print("Proceso de Mermas finalizado.")

    def obtenerURLdevoluciones(orden):
        url = "https://benny.sphinx.cl/Documento$reporteExcel.service?param={"
        parametros1 = '"obs":false,"fechaHasta":"2025-12-31","fechaDesde":"2024-01-01","idClasificacion":"10","obsProducto":"","fechaTipo":"D","promo":false,"maxRow":15,"maxPage":16,"transito":false,"pendiente":"P","detalle":true,'
        parametros2 = f'"folio":"{orden}",'
        parametros3 = '"idTipo":"50","page":1,"idGrupo":null}'
        direccion = f'{url}{parametros1}{parametros2}{parametros3}' 
        return pyperclip.copy(direccion)

    def devolucion():
        directorio = funciones.carpetaDescargas()
        archivos = funciones.buscarArchivos(directorio,".xlsx")
        for archivo in archivos:
            print(archivo)
            if "Documento" in os.path.basename(archivo):
                try:
                    workbook = openpyxl.load_workbook(archivo)
                    hojaPrincipal = workbook.active
                    numeroOrden = str(hojaPrincipal['C7'].value).rstrip('.0')
                    nuevoPDV = hojaPrincipal['L7'].value

                    # Obtener el nuevo nombre del archivo desde las celdas C7 y A4
                    newName = f"{numeroOrden} {nuevoPDV}"

                    # Hojas
                    hojaSphinx = workbook['sphinx']
                    codigos = list({cell.value for row in hojaSphinx['AY7':'AY1000'] for cell in row if cell.value is not None})

                    # Colores
                    colorAmarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    colorVioleta = PatternFill(start_color="D0CEFF", end_color="D0CEFF", fill_type="solid")
                    colorVerde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                    # Bordes
                    bordeFino = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    try:
                        hojaVerificacion = workbook.create_sheet("Verificación")
                        hojaVerificacion['A1'] = 'Cantidad en Guia:'
                        hojaVerificacion['A1'].fill = colorAmarillo
                        hojaVerificacion['B1'].value = '+SUMA(sphinx!AW:AW)'
                        hojaVerificacion['B1'].border = bordeFino
                        hojaVerificacion['A2'] = 'Cantidad pickeada:'
                        hojaVerificacion['A2'].fill = colorAmarillo
                        hojaVerificacion['B2'].value = '+SUMA(H:H)'
                        hojaVerificacion['B2'].border = bordeFino
                        
                        hojaVerificacion['A4'] = 'Código'
                        hojaVerificacion['A4'].fill = colorVioleta

                        filaV = 5
                        for codigo in codigos:
                            celdaV = hojaVerificacion.cell(row=filaV, column=1)
                            celdaV.value = codigo
                            celdaV.fill = colorVerde
                            celdaC = hojaVerificacion.cell(row=filaV, column=8)
                            celdaC.border = bordeFino
                            filaV += 1
                        hojaVerificacion['B4'] = 'Envio de PDV'
                        hojaVerificacion['B4'].fill = colorVioleta
                        hojaVerificacion['B5'].value = '+SUMAR.SI.CONJUNTO(sphinx!AW:AW;sphinx!AY:AY;A5)'
                        hojaVerificacion['C4'] = 'Recepcionado'
                        hojaVerificacion['C4'].fill = colorVioleta
                        hojaVerificacion['C5'].value = '+SUMAR.SI.CONJUNTO(I:I;H:H;A5)'
                        hojaVerificacion['D4'] = 'Estado'
                        hojaVerificacion['D4'].fill = colorVioleta
                        hojaVerificacion['D5'].value = '+SI(B5=C5;"100%";SI(B5>C5;"Falta";SI(B5<C5;"Sobra";"")))'

                        hojaVerificacion.column_dimensions['A'].width = 30
                        hojaVerificacion.column_dimensions['H'].width = 30
                    except:
                        print("No se pudo realizar la hoja de VERIFICACIONES")

                    try:
                        # Creando la seccion conteo
                        hojaVerificacion['H4'] = 'Código'
                        hojaVerificacion['H4'].fill = colorVioleta
                        hojaVerificacion['I4'] = 'Cantidad'
                        hojaVerificacion['I4'].fill = colorVioleta
                        hojaVerificacion['I5'].value = '+SI(H5=0;0;1)'
                        hojaVerificacion['J4'] = 'Estado'
                        hojaVerificacion['J4'].fill = colorVioleta
                        hojaVerificacion['J5'].value = '+SI.ERROR(SI(BUSCARV(H5;A:A;1;0)=H5;"Está");"No se encuentra en Guia")'
                    except:
                        print("No se pudo realizar la hoja de CONTEO")

                    workbook.save(os.path.join(directorio, newName + ".xlsx"))
                    print("Archivo creado")
                    os.remove(archivo)
                except:
                    print("No hay archivos Excel")
        return

if __name__ == "__main__":
    import importlib as ilib
    import subprocess as sub
    def libSetup(lib):
        # Funcion para instalar automaticamente librerias no existentes
        try:ilib.import_module(lib)
        except ImportError:sub.check_call(['pip', 'install', lib])
        return

    import os, time, csv, pyperclip
    import datetime as dt
    import base64 as b6
    from multiprocessing import Pool
    libSetup('tkinter')
    import tkinter as tk
    from tkinter import Tk, Button, Label, Text
    from tkinter import messagebox
    libSetup('warnings')
    import warnings
    libSetup('pandas')
    import pandas as pd
    libSetup('openpyxl')
    from openpyxl.styles import PatternFill , Border, Side
    libSetup('python-dotenv')
    from dotenv import load_dotenv, set_key
    libSetup('requests')
    import requests as rq
    libSetup('bs4')
    from bs4 import BeautifulSoup as bs
    libSetup('selenium')
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.support.ui import WebDriverWait, Select
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import TimeoutException, NoSuchElementException

    ventanaPrincipal = VentanaPrincipal()
    funciones.creacionEntorno()
    ventanaPrincipal.iniciar()
